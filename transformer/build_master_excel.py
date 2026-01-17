"""
Comprehensive Master Excel Builder for FADA ETL Pipeline
Extracts ALL table data from PDFs and consolidates into structured master file.

Data Categories Extracted:
- Inventory Days
- Retail Data (2W, 3W, PV, CV, TRAC, CE with all OEMs)
- EV OEMs (Two-Wheeler, Three-Wheeler, Commercial Vehicle, PV)
- Retail Strength Index (Urban and Rural RTOs)
- Motor Vehicle Road Tax Collection
- EV Penetration
- Fuel Wise Vehicle Retail Market Share
- 3W Subcategories

Structure:
- FY (yearly) data columns first
- Monthly data columns after (chronologically sorted)
"""

import re
import pandas as pd
from pathlib import Path
from datetime import datetime
from collections import defaultdict, OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from typing import List, Dict, Optional, Tuple, Any

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from config import EXCEL_DIR, OUTPUT_DIR, CACHE_FILE
from utils.logger import get_logger
from filters.date_filter import parse_month_year_from_filename


# ============== TABLE IDENTIFICATION ==============

# Table type patterns for identification
TABLE_PATTERNS = {
    # OEM Tables
    'Two Wheeler (2W)': ['two wheeler oem', 'two-wheeler oem', '2w oem'],
    'Two-Wheeler EV OEM': ['two wheeler ev', 'two-wheeler ev', 'electric two-wheeler', '2w ev'],
    'Three Wheeler (3W)': ['three wheeler oem', 'three-wheeler oem', '3w oem'],
    'Three-Wheeler EV OEM': ['three wheeler ev', 'three-wheeler ev', 'electric three-wheeler', '3w ev'],
    'Passenger Vehicle (PV)': ['passenger vehicle oem', 'pv oem', 'passenger car'],
    'PV EV OEM': ['pv ev', 'passenger vehicle ev', 'electric passenger'],
    'Tractor (TRAC)': ['tractor oem', 'tractor'],
    'Commercial Vehicle (CV)': ['commercial vehicle oem', 'cv oem'],
    'Commercial Vehicle EV OEM': ['commercial vehicle ev', 'cv ev', 'electric commercial'],
    'Construction Equipment': ['construction equipment', 'ce oem'],
    
    # Summary/Category Tables
    'Retail Data Summary': ['category', 'retail data'],
    'Inventory Days': ['inventory days', 'inventory'],
    
    # Additional Tables
    'Retail Strength Urban': ['retail strength', 'urban rto'],
    'Retail Strength Rural': ['retail strength', 'rural rto'],
    'Road Tax Collection': ['road tax', 'motor vehicle road tax'],
    'EV Penetration': ['ev penetration', 'electric vehicle penetration'],
    'Fuel Wise Market Share': ['fuel wise', 'fuel type', 'market share'],
    
    # 3W Subcategories
    '3W Subcategories': ['three-wheeler (passenger)', 'three-wheeler (goods)', 'e-rickshaw'],
}

# All row labels we want to capture
MASTER_ROW_LABELS = [
    # Summary Categories
    "2W", "3W", "PV", "CV", "TRAC", "CE", "TOTAL", "LCV", "HCV", "MCV", "OTHERS",
    
    # 3W Subcategories
    "THREE-WHEELER (PASSENGER)", "E-RICKSHAW(P)", "THREE-WHEELER (GOODS)", 
    "E-RICKSHAW WITH CART (G)", "THREE-WHEELER (PERSONAL)",
    
    # Inventory
    "INVENTORY DAYS",
    
    # All OEM names will be dynamically captured
]


def clean_value(val) -> Optional[int]:
    """Clean and convert cell values to integers."""
    if pd.isna(val) or val is None:
        return None
    if isinstance(val, (int, float)):
        if pd.notna(val):
            return int(val)
        return None
    if isinstance(val, str):
        val_clean = re.sub(r'[,\s]', '', val.strip())
        if re.match(r'^-?\d+$', val_clean):
            return int(val_clean)
    return None


def normalize_name(name: str) -> str:
    """Normalize row names for matching."""
    if pd.isna(name) or name is None:
        return ""
    name = str(name).strip()
    # Remove extra whitespace
    name = re.sub(r'\s+', ' ', name)
    return name.upper()


def extract_timepoints_from_header(header_row) -> Dict[str, int]:
    """Extract all time-based columns from header row."""
    timepoints = {}
    
    for col_idx, cell_value in enumerate(header_row):
        if pd.isna(cell_value):
            continue
        
        cell_str = str(cell_value).upper().strip()
        
        # Skip non-date columns
        if any(skip in cell_str for skip in ['%', 'YOY', 'MOM', 'GROWTH', 'MARKET SHARE', 
                                              'OEM', 'NAME', 'CATEGORY', 'TOTAL']):
            continue
        
        # Match FY patterns: FY'24, FY24, FY 2024
        fy_match = re.search(r"FY[\s']*(\d{2,4})", cell_str)
        if fy_match:
            year = fy_match.group(1)
            if len(year) == 4:
                year = year[2:]
            timepoint = f"FY'{year}"
            timepoints[timepoint] = col_idx
            continue
        
        # Match month patterns: JAN'24, JANUARY 2024, etc.
        month_match = re.search(
            r'(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)[A-Z]*[\'\s\-]*(\d{2,4})',
            cell_str
        )
        if month_match:
            month = month_match.group(1)[:3]
            year = month_match.group(2)
            if len(year) == 4:
                year = year[2:]
            timepoint = f"{month}'{year}"
            timepoints[timepoint] = col_idx
    
    return timepoints


def identify_table_type(df: pd.DataFrame, sheet_name: str = "") -> Optional[str]:
    """Identify table type from sheet name or content."""
    search_text = sheet_name.lower()
    
    # Check first 5 rows for table type indicators
    for i in range(min(5, len(df))):
        row_text = ' '.join(str(x).lower() for x in df.iloc[i].fillna(''))
        search_text += ' ' + row_text
    
    # Match against patterns
    for table_type, patterns in TABLE_PATTERNS.items():
        for pattern in patterns:
            if pattern in search_text:
                return table_type
    
    return None


def find_header_row(df: pd.DataFrame) -> Optional[int]:
    """Find the header row containing timepoints."""
    for i in range(min(10, len(df))):
        row = df.iloc[i]
        timepoints = extract_timepoints_from_header(row)
        if len(timepoints) >= 1:  # At least one timepoint found
            return i
    return None


def extract_table_data(df: pd.DataFrame) -> Tuple[str, Dict[str, Dict[str, Any]]]:
    """
    Extract all data from a table.
    Returns: (table_type, {row_label: {timepoint: value}})
    
    UPDATED: Now skips italic sub-rows (indented child items under main OEM headings).
    Only includes main summary/heading rows for cleaner output.
    """
    header_idx = find_header_row(df)
    if header_idx is None:
        return None, {}
    
    header_row = df.iloc[header_idx]
    timepoints = extract_timepoints_from_header(header_row)
    
    if not timepoints:
        return None, {}
    
    table_type = identify_table_type(df)
    data = {}
    
    # Track previous main OEM for sub-item detection
    # Sub-items typically follow their parent and have indentation or similar naming
    
    # Extract data from rows after header
    for i in range(header_idx + 1, len(df)):
        row = df.iloc[i]
        
        # Get row label from first column
        label_raw = row.iloc[0]
        if pd.isna(label_raw):
            continue
        
        label_str = str(label_raw)
        label = normalize_name(label_raw)
        if not label or label in ['', 'NAN', 'NONE']:
            continue
        
        # Skip header-like rows
        if any(skip in label for skip in ['OEM NAME', 'CATEGORY', 'SR NO', 'S.NO']):
            continue
        
        # ============== SKIP SUB-ROWS (Italic/Indented Items) ==============
        # Sub-rows are typically:
        # 1. Start with leading spaces/tabs (indentation)
        # 2. Start with numbers (like "1.", "2.")
        # 3. Have the same name as parent but in different case
        # 4. Start with lowercase
        
        # Check for leading whitespace (indentation = sub-item)
        if label_str.startswith(' ') or label_str.startswith('\t'):
            continue
        
        # Check for numbered sub-items (e.g., "1. Model Name")
        if re.match(r'^\d+\.?\s', label_str.strip()):
            continue
        
        # Skip if the row appears to be a sub-brand/model under a main OEM
        # These often have specific patterns like "XYZ - Model" when parent is "XYZ"
        # Or they are in ALL CAPS but with a sub-pattern
        
        # Extract values for each timepoint
        row_data = {}
        has_data = False
        
        for timepoint, col_idx in timepoints.items():
            if col_idx < len(row):
                value = clean_value(row.iloc[col_idx])
                if value is not None:
                    row_data[timepoint] = value
                    has_data = True
        
        if has_data:
            # Use original case-preserved label for display
            display_label = label_str.strip()
            data[display_label] = row_data
    
    return table_type, data


def sort_timepoints_columns(timepoints: set) -> List[str]:
    """Sort timepoints: FYs first (ascending), then months (chronological)."""
    fys = [tp for tp in timepoints if tp.startswith('FY')]
    months = [tp for tp in timepoints if not tp.startswith('FY')]
    
    # Sort FYs by year
    def fy_sort(tp):
        year = int(re.sub(r'\D', '', tp) or 0)
        return year
    
    fys_sorted = sorted(fys, key=fy_sort)
    
    # Sort months chronologically
    def month_sort(tp):
        try:
            dt = datetime.strptime(tp, "%b'%y")
            return (dt.year, dt.month)
        except:
            return (9999, 12)
    
    months_sorted = sorted(months, key=month_sort)
    
    return fys_sorted + months_sorted


def build_comprehensive_master(excel_dir: Path = None,
                                output_dir: Path = None) -> Optional[Path]:
    """
    Build comprehensive master Excel extracting ALL table data.
    """
    logger = get_logger()
    
    if excel_dir is None:
        excel_dir = EXCEL_DIR
    if output_dir is None:
        output_dir = OUTPUT_DIR
    
    excel_dir = Path(excel_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    excel_files = list(excel_dir.glob("*_tables.xlsx"))
    
    if not excel_files:
        logger.warning(f"No Excel files found in {excel_dir}")
        return None
    
    logger.info(f"Building comprehensive master from {len(excel_files)} files")
    
    # Data storage: {table_type: {row_label: {timepoint: value}}}
    all_data = defaultdict(lambda: defaultdict(dict))
    all_timepoints = set()
    
    # Process each file
    for file_path in sorted(excel_files):
        logger.info(f"Processing: {file_path.name}")
        
        try:
            xls = pd.ExcelFile(file_path)
            
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                
                # Skip very small tables
                if len(df) < 2:
                    continue
                
                table_type, data = extract_table_data(df)
                
                if data:
                    # If table type not identified, use sheet name
                    if not table_type:
                        table_type = f"Sheet: {sheet_name}"
                    
                    for row_label, timepoint_values in data.items():
                        for tp, val in timepoint_values.items():
                            # Merge data (newer files overwrite older)
                            all_data[table_type][row_label][tp] = val
                            all_timepoints.add(tp)
                            
        except Exception as e:
            logger.error(f"Error processing {file_path}: {e}")
            continue
    
    if not all_data:
        logger.warning("No data extracted from any files")
        return None
    
    # Sort timepoints
    sorted_timepoints = sort_timepoints_columns(all_timepoints)
    
    logger.info(f"Extracted data from {len(all_data)} table types")
    logger.info(f"Total timepoints: {len(sorted_timepoints)}")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Data"
    
    # Define styles
    header_font = Font(bold=True)
    section_font = Font(bold=True, size=12)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row_idx = 1
    
    # Define the order of table types
    table_order = [
        'Retail Data Summary',
        'Inventory Days',
        'Two Wheeler (2W)',
        'Two-Wheeler EV OEM',
        'Three Wheeler (3W)',
        'Three-Wheeler EV OEM',
        '3W Subcategories',
        'Passenger Vehicle (PV)',
        'PV EV OEM',
        'Commercial Vehicle (CV)',
        'Commercial Vehicle EV OEM',
        'Tractor (TRAC)',
        'Construction Equipment',
        'Retail Strength Urban',
        'Retail Strength Rural',
        'Road Tax Collection',
        'EV Penetration',
        'Fuel Wise Market Share',
    ]
    
    # Add tables that are in order first
    written_tables = set()
    
    for table_type in table_order:
        if table_type in all_data:
            written_tables.add(table_type)
            table_data = all_data[table_type]
            
            # Section header
            ws.cell(row_idx, 1, table_type).font = section_font
            row_idx += 1
            
            # Column headers
            ws.cell(row_idx, 1, "Item").font = header_font
            for col_idx, tp in enumerate(sorted_timepoints, 2):
                ws.cell(row_idx, col_idx, tp).font = header_font
            row_idx += 1
            
            # Data rows - sort with TOTAL last
            row_labels = sorted(table_data.keys(), 
                              key=lambda x: (x.upper() == 'TOTAL' or x.upper() == 'TOTALS', x))
            
            for label in row_labels:
                ws.cell(row_idx, 1, label)
                for col_idx, tp in enumerate(sorted_timepoints, 2):
                    val = table_data[label].get(tp)
                    if val is not None:
                        ws.cell(row_idx, col_idx, val)
                row_idx += 1
            
            row_idx += 1  # Blank row between sections
    
    # Add any remaining tables not in the predefined order
    for table_type in sorted(all_data.keys()):
        if table_type in written_tables:
            continue
        
        table_data = all_data[table_type]
        
        # Section header
        ws.cell(row_idx, 1, table_type).font = section_font
        row_idx += 1
        
        # Column headers
        ws.cell(row_idx, 1, "Item").font = header_font
        for col_idx, tp in enumerate(sorted_timepoints, 2):
            ws.cell(row_idx, col_idx, tp).font = header_font
        row_idx += 1
        
        # Data rows
        row_labels = sorted(table_data.keys(), 
                          key=lambda x: (x.upper() == 'TOTAL' or x.upper() == 'TOTALS', x))
        
        for label in row_labels:
            ws.cell(row_idx, 1, label)
            for col_idx, tp in enumerate(sorted_timepoints, 2):
                val = table_data[label].get(tp)
                if val is not None:
                    ws.cell(row_idx, col_idx, val)
            row_idx += 1
        
        row_idx += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 45
    for col_idx in range(2, len(sorted_timepoints) + 2):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = 12
    
    # Save
    output_path = output_dir / "Master_FADA_Data.xlsx"
    wb.save(output_path)
    
    # Log summary
    total_rows = sum(len(v) for v in all_data.values())
    logger.info(f"Master Excel saved to {output_path}")
    logger.info(f"  - Table types: {len(all_data)}")
    logger.info(f"  - Total row items: {total_rows}")
    logger.info(f"  - Timepoint columns: {len(sorted_timepoints)}")
    
    return output_path


def build_master_excel_for_month(month: int, year: int, 
                                  excel_dir: Path = None,
                                  output_dir: Path = None) -> Optional[Path]:
    """Build master for specific month (wrapper for API compatibility)."""
    return build_comprehensive_master(excel_dir, output_dir)


def build_consolidated_master(months: list = None, years: list = None,
                               excel_dir: Path = None,
                               output_dir: Path = None) -> Optional[Path]:
    """
    Build consolidated master Excel for specific months and years.
    
    UPDATED: Now filters Excel files based on user-selected months and years.
    If months/years are None, processes all available files.
    
    Args:
        months: List of months (1-12) to include
        years: List of years to include
        excel_dir: Directory containing extracted Excel files
        output_dir: Directory to save master file
        
    Returns:
        Path to generated master Excel file, or None if no data
    """
    logger = get_logger()
    
    if excel_dir is None:
        excel_dir = EXCEL_DIR
    if output_dir is None:
        output_dir = OUTPUT_DIR
    
    excel_dir = Path(excel_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Get all Excel files
    all_excel_files = list(excel_dir.glob("*_tables.xlsx"))
    
    if not all_excel_files:
        logger.warning(f"No Excel files found in {excel_dir}")
        return None
    
    # Filter files by months and years if specified
    if months is not None and years is not None:
        filtered_files = []
        month_names = {
            1: 'jan', 2: 'feb', 3: 'mar', 4: 'apr',
            5: 'may', 6: 'jun', 7: 'jul', 8: 'aug',
            9: 'sep', 10: 'oct', 11: 'nov', 12: 'dec'
        }
        
        for file_path in all_excel_files:
            filename = file_path.stem.lower()
            
            # Check if file matches any of the selected month/year combinations
            for year in years:
                for month in months:
                    month_name = month_names.get(month, '')
                    year_str = str(year)
                    year_short = year_str[-2:]
                    
                    # Match patterns like "jan_2025", "january_25", "01_2025", etc.
                    if (month_name in filename and (year_str in filename or year_short in filename)):
                        if file_path not in filtered_files:
                            filtered_files.append(file_path)
                            logger.info(f"Including file for {month}/{year}: {file_path.name}")
                        break
        
        if filtered_files:
            excel_files = filtered_files
            logger.info(f"Filtered to {len(excel_files)} files for selected periods")
        else:
            # If no matches found, use all files (fallback)
            logger.warning(f"No files matched selected periods, using all {len(all_excel_files)} files")
            excel_files = all_excel_files
    else:
        excel_files = all_excel_files
    
    logger.info(f"Building consolidated master from {len(excel_files)} files")
    
    # Data storage: {table_type: {row_label: {timepoint: value}}}
    all_data = defaultdict(lambda: defaultdict(dict))
    all_timepoints = set()
    
    # Process each file
    for file_path in sorted(excel_files):
        logger.info(f"Processing: {file_path.name}")
        
        try:
            xls = pd.ExcelFile(file_path)
            
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                
                # Skip very small tables
                if len(df) < 2:
                    continue
                
                table_type, data = extract_table_data(df)
                
                if data:
                    # If table type not identified, use sheet name
                    if not table_type:
                        table_type = f"Sheet: {sheet_name}"
                    
                    for row_label, timepoint_values in data.items():
                        for tp, val in timepoint_values.items():
                            # Merge data (newer files overwrite older)
                            all_data[table_type][row_label][tp] = val
                            all_timepoints.add(tp)
                            
        except Exception as e:
            logger.error(f"Error processing {file_path}: {e}")
            continue
    
    if not all_data:
        logger.warning("No data extracted from any files")
        return None
    
    # Sort timepoints
    sorted_timepoints = sort_timepoints_columns(all_timepoints)
    
    logger.info(f"Extracted data from {len(all_data)} table types")
    logger.info(f"Total timepoints: {len(sorted_timepoints)}")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Data"
    
    # Define styles
    header_font = Font(bold=True)
    section_font = Font(bold=True, size=12)
    
    row_idx = 1
    
    # Define the order of table types
    table_order = [
        'Retail Data Summary',
        'Inventory Days',
        'Two Wheeler (2W)',
        'Two-Wheeler EV OEM',
        'Three Wheeler (3W)',
        'Three-Wheeler EV OEM',
        '3W Subcategories',
        'Passenger Vehicle (PV)',
        'PV EV OEM',
        'Commercial Vehicle (CV)',
        'Commercial Vehicle EV OEM',
        'Tractor (TRAC)',
        'Construction Equipment',
        'Retail Strength Urban',
        'Retail Strength Rural',
        'Road Tax Collection',
        'EV Penetration',
        'Fuel Wise Market Share',
    ]
    
    # Add tables that are in order first
    written_tables = set()
    
    for table_type in table_order:
        if table_type in all_data:
            written_tables.add(table_type)
            table_data = all_data[table_type]
            
            # Section header
            ws.cell(row_idx, 1, table_type).font = section_font
            row_idx += 1
            
            # Column headers
            ws.cell(row_idx, 1, "Item").font = header_font
            for col_idx, tp in enumerate(sorted_timepoints, 2):
                ws.cell(row_idx, col_idx, tp).font = header_font
            row_idx += 1
            
            # Data rows - sort with TOTAL last
            row_labels = sorted(table_data.keys(), 
                              key=lambda x: (x.upper() == 'TOTAL' or x.upper() == 'TOTALS', x))
            
            for label in row_labels:
                ws.cell(row_idx, 1, label)
                for col_idx, tp in enumerate(sorted_timepoints, 2):
                    val = table_data[label].get(tp)
                    if val is not None:
                        ws.cell(row_idx, col_idx, val)
                row_idx += 1
            
            row_idx += 1  # Blank row between sections
    
    # Add any remaining tables not in the predefined order
    for table_type in sorted(all_data.keys()):
        if table_type in written_tables:
            continue
        
        table_data = all_data[table_type]
        
        # Section header
        ws.cell(row_idx, 1, table_type).font = section_font
        row_idx += 1
        
        # Column headers
        ws.cell(row_idx, 1, "Item").font = header_font
        for col_idx, tp in enumerate(sorted_timepoints, 2):
            ws.cell(row_idx, col_idx, tp).font = header_font
        row_idx += 1
        
        # Data rows
        row_labels = sorted(table_data.keys(), 
                          key=lambda x: (x.upper() == 'TOTAL' or x.upper() == 'TOTALS', x))
        
        for label in row_labels:
            ws.cell(row_idx, 1, label)
            for col_idx, tp in enumerate(sorted_timepoints, 2):
                val = table_data[label].get(tp)
                if val is not None:
                    ws.cell(row_idx, col_idx, val)
            row_idx += 1
        
        row_idx += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 45
    for col_idx in range(2, len(sorted_timepoints) + 2):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = 12
    
    # Generate filename with period info
    if months and years:
        period_str = f"_{min(years)}_{max(years)}"
        output_filename = f"Master_FADA_Data{period_str}.xlsx"
    else:
        output_filename = "Master_FADA_Data.xlsx"
    
    output_path = output_dir / output_filename
    wb.save(output_path)
    
    # Log summary
    total_rows = sum(len(v) for v in all_data.values())
    logger.info(f"Master Excel saved to {output_path}")
    logger.info(f"  - Table types: {len(all_data)}")
    logger.info(f"  - Total row items: {total_rows}")
    logger.info(f"  - Timepoint columns: {len(sorted_timepoints)}")
    
    return output_path


if __name__ == '__main__':
    result = build_comprehensive_master()
    if result:
        print(f"Success! Created: {result}")
    else:
        print("No data to consolidate.")

